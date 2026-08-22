"""Generate three intentionally different monthly finance packages."""

from __future__ import annotations

import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.paths import RAW_DIR

MONTHS = pd.date_range("2025-04-30", "2026-06-30", freq="ME")
SCENARIOS = ["Actual", "Budget", "Forecast", "Prior Year"]
SEED = 20260630


def _scenario_value(actual: float, scenario: str, month_index: int, rng: np.random.Generator) -> float:
    """Create a realistic comparison scenario around an actual value."""
    if scenario == "Actual":
        factor = 1.0
    elif scenario == "Budget":
        factor = 0.985 + 0.008 * math.sin(month_index / 2)
    elif scenario == "Forecast":
        factor = 0.995 + 0.006 * math.cos(month_index / 3)
    else:
        factor = 0.93 + 0.01 * math.sin(month_index / 3)
    return round(actual * factor * rng.normal(1.0, 0.006), 3)


def _commercial_banking(rng: np.random.Generator) -> pd.DataFrame:
    """Build a stacked long-format Commercial Banking package."""
    base = {
        "Loan Interest": 5.60,
        "Treasury Fee": 2.35,
        "Deposit Fee": 0.92,
        "Merchant Fee": 1.18,
        "FX Fee": 0.48,
        "Operating Expense": 3.85,
        "Loan Balance": 1650.0,
        "Deposit Balance": 1475.0,
    }
    rows: list[dict] = []
    for i, period in enumerate(MONTHS):
        for metric, start in base.items():
            growth = 1 + (0.0065 if "Balance" in metric else 0.0045) * i
            seasonality = 1 + 0.025 * math.sin((i + 1) * math.pi / 6)
            actual = start * growth * seasonality
            if period >= pd.Timestamp("2026-04-30") and metric in {"Loan Interest", "Loan Balance"}:
                actual *= 1 + 0.018 * (i - 11)
            for scenario in SCENARIOS:
                amount = _scenario_value(actual, scenario, i, rng)
                if scenario == "Budget" and metric == "Operating Expense":
                    amount = round(actual / 1.035, 3)
                rows.append(
                    {
                        "Date": period.strftime("%m/%d/%Y"),
                        "Metric": metric,
                        "Scenario": scenario,
                        "Amount": amount,
                        "Unit": "USD millions",
                    }
                )
    return pd.DataFrame(rows)


def _commercial_real_estate(rng: np.random.Generator) -> pd.DataFrame:
    """Build a wide CRE package with one row per month and scenario."""
    base = {
        "NII": 6.35,
        "Orig Fee": 0.78,
        "Prepay Fee": 0.22,
        "Servicing Fee": 0.34,
        "Other Fee": 0.16,
        "Credit Loss Provision": 0.46,
        "Opex": 2.15,
        "CRE Loan Bal ($MM)": 1900.0,
        "NPL Proxy ($MM)": 20.0,
    }
    rows: list[dict] = []
    for i, period in enumerate(MONTHS):
        actuals: dict[str, float] = {}
        for metric, start in base.items():
            growth = 1 + (0.004 if "Loan Bal" in metric else 0.0025) * i
            actuals[metric] = start * growth * (1 + 0.018 * math.sin(i * math.pi / 5))
        if period >= pd.Timestamp("2026-04-30"):
            quarter_step = i - 11
            actuals["Orig Fee"] *= 1 + 0.06 * quarter_step
            actuals["CRE Loan Bal ($MM)"] *= 1 + 0.012 * quarter_step
            actuals["Credit Loss Provision"] *= 1 + 0.10 * quarter_step
            actuals["NPL Proxy ($MM)"] *= 1 + 0.045 * quarter_step
        for scenario in SCENARIOS:
            row = {"Period": period.strftime("%Y-%m"), "Case": scenario}
            for metric, actual in actuals.items():
                amount = _scenario_value(actual, scenario, i, rng)
                if scenario == "Forecast" and metric == "Credit Loss Provision" and period >= pd.Timestamp("2026-05-31"):
                    amount = round(actual / 1.10, 3)
                row[metric] = amount
            rows.append(row)
    return pd.DataFrame(rows)


def _capital_markets(rng: np.random.Generator) -> dict[str, pd.DataFrame]:
    """Build one transposed month-across-columns sheet per Capital Markets metric."""
    base = {
        "Advisory Fee": 2.10,
        "Underwriting Revenue": 2.45,
        "Trading Revenue": 1.85,
        "Structuring Fee": 0.72,
        "Syndication Fee": 0.60,
        "Operating Expense": 3.35,
    }
    sheets: dict[str, pd.DataFrame] = {}
    for metric, start in base.items():
        records = []
        for scenario in SCENARIOS:
            row = {"Scenario": scenario}
            for i, period in enumerate(MONTHS):
                actual = start * (1 + 0.0015 * i) * (1 + 0.08 * math.sin((i + 2) * math.pi / 4))
                if period == pd.Timestamp("2026-06-30"):
                    boost = 1.42 if metric != "Operating Expense" else 1.16
                    actual *= boost
                amount = _scenario_value(actual, scenario, i, rng)
                if scenario in {"Budget", "Forecast"} and period == pd.Timestamp("2026-06-30"):
                    divisor = 1.18 if metric != "Operating Expense" else 1.08
                    amount = round(actual / divisor, 3)
                if scenario == "Prior Year" and period == pd.Timestamp("2026-06-30") and metric != "Operating Expense":
                    # Public Q2 2026 capital markets revenue was 62.5% above Q2 2025.
                    amount = round(actual / 1.625, 3)
                row[period.strftime("%b-%y")] = amount
            records.append(row)
        sheets[metric] = pd.DataFrame(records)
    return sheets


def _style_workbook(path: Path) -> None:
    """Apply a compact finance-package style to an existing workbook."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    navy = "17365D"
    light_blue = "D9EAF7"
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in sheet.iter_rows(min_row=2):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=light_blue)
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 24)
            sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(path)


def generate_raw_packages(output_dir: Path = RAW_DIR) -> list[Path]:
    """Generate and save all three source workbooks."""
    output_dir.mkdir(parents=True, exist_ok=True)
    rng = np.random.default_rng(SEED)
    paths = [
        output_dir / "commercial_banking_monthly.xlsx",
        output_dir / "commercial_real_estate_monthly.xlsx",
        output_dir / "capital_markets_monthly.xlsx",
    ]

    _commercial_banking(rng).to_excel(paths[0], index=False, sheet_name="Monthly Detail")
    _commercial_real_estate(rng).to_excel(paths[1], index=False, sheet_name="CRE Monthly")
    with pd.ExcelWriter(paths[2], engine="openpyxl") as writer:
        for metric, frame in _capital_markets(rng).items():
            frame.to_excel(writer, index=False, sheet_name=metric[:31])
    for path in paths:
        _style_workbook(path)
    return paths


if __name__ == "__main__":
    for generated_path in generate_raw_packages():
        print(f"Generated {generated_path}")
